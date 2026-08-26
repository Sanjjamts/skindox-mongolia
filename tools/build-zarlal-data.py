#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel -> rtqfy9q7pd-data.js
=======================
Гишүүнчлэл цуцлагдах эрсдэлтэй гишүүдийн жагсаалтыг Excel файлаас уншиж,
rtqfy9q7pd.html хуудсанд хэрэглэгддэг `rtqfy9q7pd-data.js` файлыг үүсгэнэ.

Хэрэглээ:
    py tools/build-zarlal-data.py "C:\\path\\to\\zarlah member_new.xlsx"

Excel-ийн шаардлага — 1-р мөр толгой, багана нь дарааллаараа:
    A: Гишүүний ID        B: Гишүүний нэр       C: Гишүүнээр элссэн огноо
    D: Гишүүний статус    E: Шууд уригч ID      F: Шууд уригчийн нэр
    G: Спонсор ID         H: Спонсор нэр        I: Уригчийн уригч ID
    J: Уригчийн уригч нэр
Sheet-ийн нэр ямар ч байж болно (эхний sheet-ийг, эсвэл "Зарлах"/"Уригчтайгаа"
нэртэйг нь автоматаар олно). Хоосон буюу "#N/A" утгыг холбоосгүй гэж үзнэ.

Дараа нь:
    git add rtqfy9q7pd-data.js
    git commit -m "Зарлал шинэчлэв"
    git push origin main
"""

import datetime
import io
import json
import os
import re
import sys

import openpyxl

PREFERRED_SHEETS = ("Зарлах", "Уригчтайгаа", "Зарлах жагсаалт")
MISSING = ("", "#N/A", "#N/A!", "#VALUE!", "#REF!", "NONE", "NULL")

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "rtqfy9q7pd-data.js")


def clean(v):
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def missing(v):
    return v.upper() in MISSING


def pick_sheet(wb):
    for name in PREFERRED_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    return wb.worksheets[0]


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = sys.argv[1]

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = pick_sheet(wb)
    print('Sheet: "%s"' % ws.title)

    people, index, rows = [], {}, []

    def ref(pid, pname):
        """Return the index of this person in the shared pool, or -1 if unknown."""
        if missing(pid):
            return -1
        if pid not in index:
            index[pid] = len(people)
            people.append([pid, pname])
        elif not people[index[pid]][1] and pname:
            people[index[pid]][1] = pname
        return index[pid]

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        c = [clean(x) for x in r[:10]] + [""] * (10 - len(r[:10]))
        mid, name, date, _status, inv_id, inv_nm, spon_id, spon_nm, gp_id, gp_nm = c[:10]
        rows.append([mid, name, date,
                     ref(inv_id, inv_nm),
                     ref(spon_id, spon_nm),
                     ref(gp_id, gp_nm)])

    dupes = len(rows) - len(set(r[0] for r in rows))
    if dupes:
        print("АНХААР: %d давхардсан гишүүний ID байна." % dupes)

    data = {
        "updated": datetime.date.today().isoformat(),
        "total": len(rows),
        "people": people,
        "rows": rows,
    }
    js = ("/* SKINDOX - Зарлах жагсаалт (auto-generated, do not edit by hand) */\n"
          "window.SKINDOX_ZARLAL=" + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    with io.open(OUT, "w", encoding="utf-8") as f:
        f.write(js)

    def linked(i):
        return sum(1 for r in rows if r[i] >= 0)

    print("OK -> %s (%d bytes)" % (OUT, os.path.getsize(OUT)))
    print("   гишүүн: %d | нэрсийн сан: %d" % (len(rows), len(people)))
    print("   уригчтай: %d | спонсортой: %d | уригчийн уригчтай: %d"
          % (linked(3), linked(4), linked(5)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
