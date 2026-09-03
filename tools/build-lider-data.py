#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel -> nkekkpn5wc-data.js
===========================
Цолтой (Polaris / Royal Polaris / Crown Polaris) гишүүдийн эрсдэлийн
жагсаалтыг Excel-ээс уншиж, nkekkpn5wc.html хуудасны өгөгдлийг үүсгэнэ.

Хэрэглээ:
    py tools/build-lider-data.py "C:\\path\\to\\lider.xlsx"

Excel-ийн шаардлага — эхний sheet дээр толгойн мөр нь "ID", "Name" гэсэн
багана агуулсан байх (ихэвчлэн 6-р мөр; дээр нь нэгтгэлийн хүснэгт байж
болно). Багануудыг ДУГААРААР нь биш НЭРЭЭР нь олдог тул шинэ багана
дунд нь нэмэгдсэн ч ажиллана. Хэрэглэдэг баганууд:
    ID · Name · Status · Идэвхи · Өөрийн ХА дүн
    Уригчийн ID · Уригч · Спонсрын ID · Спонсор
    Уригчийн уригч ID · Уригчийн уригч нэр
    Уригч, уригчийн уригч ROYAL,CROWN · Шууд уригч Polaris
Хоосон буюу "#N/A" утгыг холбоосгүй гэж үзнэ.

Дараа нь:
    git add nkekkpn5wc-data.js
    git commit -m "Лидерийн зарлал шинэчлэв"
    git push origin main
"""

import collections
import datetime
import io
import json
import os
import re
import sys

import openpyxl

MISSING = ("", "#N/A", "#N/A!", "#VALUE!", "#REF!", "NONE", "NULL")

# Гаралтын талбар -> Excel-ийн толгойн нэр (жижиг үсгээр, хэсэгчилсэн тааруулалт)
COLUMNS = [
    ("id",        "id"),
    ("name",      "name"),
    ("rank",      "status"),
    ("activity",  "идэвхи"),
    ("ownpv",     "өөрийн ха дүн"),
    ("inv_id",    "уригчийн id"),
    ("inv_name",  "уригч"),
    ("spon_id",   "спонсрын id"),
    ("spon_name", "спонсор"),
    ("up_id",     "уригчийн уригч id"),
    ("up_name",   "уригчийн уригч нэр"),
    ("royal",     "уригч, уригчийн уригч royal"),
    ("polaris",   "шууд уригч polaris"),
]

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "nkekkpn5wc-data.js")


def clean(v):
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def blank(v):
    return v.upper() in MISSING


def num(v):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def find_header(ws):
    """Return (row index, {header text -> column index}) for the data header row."""
    for row in ws.iter_rows(min_row=1, max_row=30):
        cells = {clean(c.value).lower(): c.column - 1 for c in row if c.value is not None}
        if "id" in cells and "name" in cells:
            return row[0].row, cells
    raise SystemExit('ERROR: "ID" болон "Name" багана бүхий толгойн мөр олдсонгүй.')


def resolve(headers):
    """Map each output field to a column index, longest header match first."""
    out = {}
    for field, want in COLUMNS:
        hit = None
        for text, idx in headers.items():
            if text == want:
                hit = idx
                break
            if text.startswith(want) and hit is None:
                hit = idx
        out[field] = hit
    for field in ("id", "name", "rank", "activity"):
        if out[field] is None:
            raise SystemExit('ERROR: "%s" талбарт тохирох багана олдсонгүй.' % field)
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = sys.argv[1]

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.worksheets[0]
    header_row, headers = find_header(ws)
    col = resolve(headers)
    print('Sheet: "%s" | толгойн мөр: %d' % (ws.title, header_row))

    def get(row, field):
        i = col[field]
        return clean(row[i]) if i is not None and i < len(row) else ""

    def link(row, id_field, name_field):
        pid = get(row, id_field)
        return ("", "") if blank(pid) else (pid, get(row, name_field))

    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or col["id"] >= len(r) or not r[col["id"]]:
            continue
        inv = link(r, "inv_id", "inv_name")
        spon = link(r, "spon_id", "spon_name")
        up = link(r, "up_id", "up_name")
        own = col["ownpv"]
        rows.append([
            get(r, "id"), get(r, "name"), get(r, "rank"), get(r, "activity"),
            num(r[own]) if own is not None and own < len(r) else 0,
            inv[0], inv[1], spon[0], spon[1], up[0], up[1],
            get(r, "royal"), get(r, "polaris"),
        ])

    dupes = len(rows) - len(set(r[0] for r in rows))
    if dupes:
        print("АНХААР: %d давхардсан ID байна." % dupes)

    data = {
        "updated": datetime.date.today().isoformat(),
        "total": len(rows),
        "rows": rows,
    }
    js = ("/* SKINDOX - Цолтой гишүүдийн зарлал (auto-generated, do not edit by hand) */\n"
          "window.SKINDOX_LIDER=" + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    with io.open(OUT, "w", encoding="utf-8") as f:
        f.write(js)

    print("OK -> %s (%d bytes, %d гишүүн)" % (OUT, os.path.getsize(OUT), len(rows)))
    print("   цол:", dict(collections.Counter(r[2] for r in rows)))
    print("   идэвхи:", dict(collections.Counter(r[3] for r in rows)))
    print("   уригчтай: %d | спонсортой: %d | уригчийн уригчтай: %d"
          % (sum(1 for r in rows if r[5]), sum(1 for r in rows if r[7]), sum(1 for r in rows if r[9])))
    return 0


if __name__ == "__main__":
    sys.exit(main())
